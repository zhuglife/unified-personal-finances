#!/usr/bin/env python3
"""Generate Financial Dashboard Excel (English)
Auto-generated script to create Financial_Dashboard.xlsx with demo data.
Requires: pandas, openpyxl
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference

data = {'Category': ['Food', 'Transport', 'Health', 'Leisure', 'Utilities', 'Other'], 'Responsible': ['Self', 'Self', 'Family Plan', 'Self', 'Admin', 'Self'], 'Amount (USD)': [1800.0, 750.0, 2200.0, 900.0, 2600.0, 400.0]}

df = pd.DataFrame(data)

wb = Workbook()
ws = wb.active
ws.title = "Financial Dashboard"

# basic styles
header_fill = PatternFill("solid", fgColor="000000")
cell_fill = PatternFill("solid", fgColor="1E1E1E")
white_font = Font(color="FFFFFF", name="Consolas", size=10)
header_font = Font(color="FFFFFF", name="Consolas", size=12, bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Border(left=Side(style="thin", color="444444"),
              right=Side(style="thin", color="444444"),
              top=Side(style="thin", color="444444"),
              bottom=Side(style="thin", color="444444"))

# write headers
headers = list(df.columns)
ws.append(headers)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin

for row in df.itertuples(index=False):
    ws.append(row)

for r in ws.iter_rows(min_row=2, max_row=1+len(df), min_col=1, max_col=3):
    for c in r:
        c.fill = cell_fill
        c.font = white_font
        c.alignment = center
        c.border = thin

# summary
initial_balance = 50000.00
total_spent = df["Amount (USD)"].sum()
remaining = initial_balance - total_spent
usage_pct = round((total_spent / initial_balance) * 100, 2)

summary_labels = ["Initial Balance", "Total Spent", "Remaining Balance", "Balance Usage (%)"]
start_col = 5
for i, lbl in enumerate(summary_labels, start=2):
    ws.cell(row=i, column=start_col, value=lbl).font = header_font
    ws.cell(row=i, column=start_col+1, value=[initial_balance, total_spent, remaining, usage_pct][i-2]).font = white_font

# charts (pie + bar)
pie = PieChart()
pie.title = "Expenses by Category"
data_ref = Reference(ws, min_col=3, min_row=1, max_row=1+len(df))
labels_ref = Reference(ws, min_col=1, min_row=2, max_row=1+len(df))
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(labels_ref)
ws.add_chart(pie, "A12")

bar = BarChart()
bar.title = "Expenses by Person"
data_ref2 = Reference(ws, min_col=3, min_row=1, max_row=1+len(df))
cats_ref2 = Reference(ws, min_col=2, min_row=2, max_row=1+len(df))
bar.add_data(data_ref2, titles_from_data=True)
bar.set_categories(cats_ref2)
ws.add_chart(bar, "I12")

wb.save("Financial_Dashboard.xlsx")
print("Dashboard generated: Financial_Dashboard.xlsx")
